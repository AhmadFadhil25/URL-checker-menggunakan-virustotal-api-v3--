"""
URL Reputation Checker - Streamlit App
Mengecek reputasi URL secara massal menggunakan VirusTotal API v3.
Jalankan dengan: streamlit run app.py
"""

import base64
import time
from io import BytesIO

import pandas as pd
import requests
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill

st.set_page_config(page_title="URL Reputation Checker", page_icon="🛡️", layout="wide")

VT_BASE_URL = "https://www.virustotal.com/api/v3"
PHISHING_STATUS = "TERINDIKASI PHISHING"
SAFE_STATUS = "TIDAK TERINDIKASI PHISHING"
ERROR_STATUSES = ("URL TIDAK VALID", "TIMEOUT", "ERROR", "URL KOSONG")


# ================= RATE LIMITER =================
class RateLimiter:
    """Menjaga jeda minimum antar request ke VirusTotal (menghindari 429)."""

    def __init__(self, interval_seconds: float):
        self.interval = interval_seconds
        self.last = 0.0

    def wait(self):
        if self.interval <= 0:
            return
        remaining = self.interval - (time.monotonic() - self.last)
        if self.last > 0 and remaining > 0:
            time.sleep(remaining)
        self.last = time.monotonic()


# ================= VIRUSTOTAL API =================
def vt_request(session, limiter, method, endpoint, **kwargs):
    limiter.wait()
    resp = session.request(method, f"{VT_BASE_URL}{endpoint}", timeout=30, **kwargs)
    if resp.status_code == 401:
        raise RuntimeError("API key VirusTotal tidak valid.")
    if resp.status_code == 429:
        raise RuntimeError("Rate limit tercapai (429). Perbesar jeda antar request lalu coba lagi.")
    return resp


def url_to_vt_id(url: str) -> str:
    return base64.urlsafe_b64encode(url.encode()).decode().rstrip("=")


def vt_get_existing_report(session, limiter, url):
    resp = vt_request(session, limiter, "GET", f"/urls/{url_to_vt_id(url)}")
    if resp.status_code == 404:
        return None
    resp.raise_for_status()
    attrs = resp.json().get("data", {}).get("attributes", {})
    last_analysis = attrs.get("last_analysis_date")
    if last_analysis is None:
        return None
    malicious = int(attrs.get("last_analysis_stats", {}).get("malicious", 0))
    age_days = (time.time() - last_analysis) / 86_400
    return malicious, age_days


def vt_submit_url(session, limiter, url):
    resp = vt_request(session, limiter, "POST", "/urls", data={"url": url})
    if resp.status_code == 400:
        raise ValueError(f"URL ditolak oleh VirusTotal: {url}")
    resp.raise_for_status()
    analysis_id = resp.json().get("data", {}).get("id")
    if not analysis_id:
        raise RuntimeError("Gagal mendapatkan Analysis ID dari VirusTotal.")
    return analysis_id


def vt_get_analysis(session, limiter, analysis_id, poll_max_attempts, log):
    for attempt in range(1, poll_max_attempts + 1):
        resp = vt_request(session, limiter, "GET", f"/analyses/{analysis_id}")
        resp.raise_for_status()
        attrs = resp.json().get("data", {}).get("attributes", {})
        status = attrs.get("status", "unknown")
        if status == "completed":
            return attrs.get("stats", {})
        log(f"Analisis belum selesai ({status}). Cek ulang {attempt}/{poll_max_attempts}...")
    log("Batas cek tercapai. Skor diasumsikan AMAN (0 malicious) sementara.")
    return {"malicious": 0}


def check_url_virustotal(session, limiter, url, max_age_days, poll_max_attempts, log):
    existing = vt_get_existing_report(session, limiter, url)
    if existing is not None:
        malicious, age_days = existing
        if age_days <= max_age_days:
            log(f"Menggunakan laporan lama ({age_days:.1f} hari lalu).")
            return malicious, f"Laporan lama ({age_days:.1f} hari)"
        log(f"Laporan berumur {age_days:.1f} hari. Melakukan scan baru...")
    else:
        log("Belum ada laporan. Melakukan scan baru...")

    analysis_id = vt_submit_url(session, limiter, url)
    stats = vt_get_analysis(session, limiter, analysis_id, poll_max_attempts, log)
    return int(stats.get("malicious", 0)), "Scan baru"


# ================= STYLING =================
def status_row_style(row):
    if row.get("Status") == PHISHING_STATUS:
        return ["background-color: #FFF2CC; font-weight: bold"] * len(row)
    if row.get("Status") in ERROR_STATUSES:
        return ["background-color: #F4CCCC; font-weight: bold"] * len(row)
    return [""] * len(row)


def to_excel_bytes(df: pd.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Hasil")
        ws = writer.sheets["Hasil"]

        fill_yellow = PatternFill("solid", fgColor="FFF2CC")
        fill_error = PatternFill("solid", fgColor="F4CCCC")
        bold = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="0B1F3A")
        for c in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        status_col = df.columns.get_loc("Status") + 1
        for r in range(2, len(df) + 2):
            status_val = ws.cell(row=r, column=status_col).value
            fill = fill_yellow if status_val == PHISHING_STATUS else (fill_error if status_val in ERROR_STATUSES else None)
            for c in range(1, len(df.columns) + 1):
                cell = ws.cell(row=r, column=c)
                if fill:
                    cell.fill = fill
                    cell.font = bold
                if c > 1:
                    cell.alignment = center

        for col_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 60)

    return output.getvalue()


# ================= UI =================
st.title("🛡️ URL Reputation Checker")
st.caption(
    "Cek reputasi banyak URL sekaligus menggunakan VirusTotal API — upload file atau tempel "
    "daftar URL, lalu unduh hasilnya sebagai Excel yang sudah diberi warna otomatis."
)

try:
    _secret_key = st.secrets.get("VT_API_KEY", "").strip()
except Exception:
    _secret_key = ""

with st.sidebar:
    st.header("⚙️ Pengaturan")
    if _secret_key:
        api_key = _secret_key
        st.success("API key sudah dikonfigurasi oleh admin (lewat Secrets). Anda tidak perlu memasukkannya.")
    else:
        api_key = st.text_input(
            "VirusTotal API Key",
            type="password",
            help="Key hanya dipakai selama sesi ini di browser Anda dan tidak disimpan di server.",
        )
    threshold = st.number_input("Ambang batas 'malicious' (di atas ini = phishing)", min_value=0, value=3, step=1)
    max_age_days = st.number_input("Toleransi umur laporan (hari)", min_value=0, value=10, step=1)
    poll_max_attempts = st.number_input("Maks. cek ulang status scan baru", min_value=1, value=3, step=1)
    request_interval = st.number_input(
        "Jeda antar request ke VirusTotal (detik)",
        min_value=0, value=16, step=1,
        help="16 detik cocok untuk akun gratis (4 request/menit). Isi 0 jika memakai akun premium.",
    )
    st.divider()
    st.caption("Dapatkan API key gratis di virustotal.com → Profil → API Key.")

tab1, tab2 = st.tabs(["📤 Upload File", "✍️ Input Manual"])

source_df = None
url_col = None

with tab1:
    uploaded = st.file_uploader("Upload file Excel (.xlsx) atau CSV berisi daftar URL", type=["xlsx", "csv"])
    if uploaded is not None:
        source_df = pd.read_csv(uploaded) if uploaded.name.endswith(".csv") else pd.read_excel(uploaded)
        st.write("Pratinjau file:")
        st.dataframe(source_df.head(), use_container_width=True)
        url_col = st.selectbox("Pilih kolom yang berisi URL", source_df.columns)
        if "Malicious" in source_df.columns and "Status" in source_df.columns:
            st.info("File ini sudah punya kolom **Malicious** & **Status** — baris yang sudah terisi akan dilewati (lanjutan dari proses sebelumnya).")

with tab2:
    manual_text = st.text_area(
        "Tempel daftar URL (satu URL per baris)", height=220,
        placeholder="https://contoh1.com\nhttps://contoh2.com",
    )
    if manual_text.strip():
        manual_urls = [u.strip() for u in manual_text.splitlines() if u.strip()]
        source_df = pd.DataFrame({"URL": manual_urls})
        url_col = "URL"

ready = source_df is not None and url_col is not None and not source_df.empty
if ready:
    st.success(f"Siap memproses **{len(source_df)}** baris data.")

run = st.button("🔍 Mulai Pengecekan", type="primary", disabled=not ready)

if "results_df" not in st.session_state:
    st.session_state.results_df = None

if run:
    if not api_key.strip():
        st.error("Masukkan VirusTotal API Key di sidebar terlebih dahulu.")
        st.stop()

    work_df = source_df.copy().reset_index(drop=True)
    work_df["URL"] = work_df[url_col].astype(str).str.strip()
    for col in ("Malicious", "Status", "Keterangan"):
        if col not in work_df.columns:
            work_df[col] = None
        # Paksa dtype object -- kolom kosong dari file Excel bisa terbaca
        # sebagai float64 oleh pandas, yang menolak diisi teks (pandas >= 2.x/3.x).
        work_df[col] = work_df[col].astype(object)

    session = requests.Session()
    session.headers.update({
        "x-apikey": api_key.strip(),
        "Accept": "application/json",
        "User-Agent": "Streamlit-URL-Checker/1.0",
    })
    limiter = RateLimiter(request_interval)

    total = len(work_df)
    to_process = [i for i in range(total) if pd.isna(work_df.loc[i, "Status"]) or str(work_df.loc[i, "Status"]).strip() == ""]
    skipped = total - len(to_process)

    progress = st.progress(0.0)
    status_placeholder = st.empty()
    log_placeholder = st.empty()
    table_placeholder = st.empty()

    if skipped:
        status_placeholder.info(f"{skipped} baris dilewati (sudah punya hasil sebelumnya).")

    # Saat proses berjalan, kolom "Keterangan" tetap ditampilkan (berguna untuk
    # memantau progres). Setelah selesai, hasil akhir & file Excel hanya berisi
    # URL, Malicious, Status.
    progress_cols = ["URL", "Malicious", "Status", "Keterangan"]
    final_cols = ["URL", "Malicious", "Status"]

    try:
        for done, idx in enumerate(to_process, start=1):
            url = work_df.loc[idx, "URL"]

            if not url or url.lower() == "nan":
                work_df.loc[idx, ["Malicious", "Status", "Keterangan"]] = [None, "URL KOSONG", "Baris kosong dilewati"]
            else:
                status_placeholder.markdown(f"**[{done}/{len(to_process)}]** Memproses: `{url}`")

                def log(msg):
                    log_placeholder.caption(msg)

                try:
                    malicious, source = check_url_virustotal(
                        session, limiter, url, max_age_days, poll_max_attempts, log
                    )
                    status = PHISHING_STATUS if malicious > threshold else SAFE_STATUS
                    work_df.loc[idx, ["Malicious", "Status", "Keterangan"]] = [malicious, status, source]
                except ValueError as e:
                    work_df.loc[idx, ["Malicious", "Status", "Keterangan"]] = [None, "URL TIDAK VALID", str(e)]
                except requests.exceptions.Timeout:
                    work_df.loc[idx, ["Malicious", "Status", "Keterangan"]] = [None, "TIMEOUT", "Koneksi timeout"]
                except requests.RequestException as e:
                    work_df.loc[idx, ["Malicious", "Status", "Keterangan"]] = [None, "ERROR", str(e)]
                except RuntimeError as e:
                    work_df.loc[idx, ["Malicious", "Status", "Keterangan"]] = [None, "ERROR", str(e)]
                    progress.progress(done / max(len(to_process), 1))
                    table_placeholder.dataframe(
                        work_df[progress_cols].style.apply(status_row_style, axis=1),
                        use_container_width=True,
                    )
                    status_placeholder.error(f"Berhenti: {e}")
                    break

            progress.progress(done / max(len(to_process), 1))
            table_placeholder.dataframe(
                work_df[progress_cols].style.apply(status_row_style, axis=1),
                use_container_width=True,
            )
    finally:
        session.close()

    st.session_state.results_df = work_df[final_cols]
    log_placeholder.empty()
    status_placeholder.success("✅ Selesai memproses semua URL.")

if st.session_state.results_df is not None:
    df_res = st.session_state.results_df
    st.subheader("📊 Hasil Pengecekan")
    st.dataframe(df_res.style.apply(status_row_style, axis=1), use_container_width=True)

    c1, c2, c3 = st.columns(3)
    c1.metric("Total URL", len(df_res))
    c2.metric("Terindikasi Phishing", int((df_res["Status"] == PHISHING_STATUS).sum()))
    c3.metric("Error / Tidak Valid", int(df_res["Status"].isin(ERROR_STATUSES).sum()))

    st.download_button(
        "⬇️ Unduh Hasil (Excel)",
        data=to_excel_bytes(df_res),
        file_name="hasil_url_checker.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
